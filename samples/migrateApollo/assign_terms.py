import argparse
import json
import os
import sys

import openpyxl

from pyapacheatlas.core import AtlasClient, AtlasEntity
from pyapacheatlas.auth import ServicePrincipalAuthentication


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file-path",
                        help="The path to the Apollo export.")
    parser.add_argument(
        "--sheetname",
        help="The name of the sheet to pull data out of for the given file-path.")
    parser.add_argument(
        "--glossary",
        help="The name of the glossary. Defaults to 'Glossary'",
        default='Glossary'
    )
    args = parser.parse_args()

    # Authenticate against your Atlas server
    oauth = ServicePrincipalAuthentication(
        tenant_id=os.environ.get("TENANT_ID", ""),
        client_id=os.environ.get("CLIENT_ID", ""),
        client_secret=os.environ.get("CLIENT_SECRET", "")
    )
    atlas_client = AtlasClient(
        endpoint_url=os.environ.get("ENDPOINT_URL", ""),
        authentication=oauth
    )

    # Read in the args.file_path
    # BEGIN EXCEL PROCESSING
    wb = openpyxl.load_workbook(args.file_path)
    required_sheet = wb.sheetnames[0]
    # Read the header of the args.sheetname
    # Otherwise, get the first sheet
    if args.sheetname:
        required_sheet = args.sheetname

    print(f"Reading {args.file_path} and sheet {args.sheetname}")

    data_sheet = wb.get_sheet_by_name(required_sheet)

    # Getting the headers
    print("Beginning reading sheet")
    sheet_headers = []
    null_threshold = 0.5
    for row in data_sheet.iter_rows(min_row=1, max_col=data_sheet.max_column, max_row=1):
        for idx, cell in enumerate(row):
            sheet_headers.append(cell.value)

    sheet_data = []
    for row in data_sheet.iter_rows(min_row=2, max_col=data_sheet.max_column,
                                    max_row=data_sheet.max_row + 1):
        row_data = {}
        nulls = 0
        for idx, cell in enumerate(row):
            row_data[sheet_headers[idx]] = (cell.value or "").strip()
            if cell.value is None:
                nulls = nulls + 1

        if float(nulls) / float(len(row_data)) < null_threshold:
            sheet_data.append(row_data)

    # END EXCEL PROCESSING

    # Extract the following fields:
    # key_headers = [
    #     "fld_it_name", "fld_business_name", "fld_it_type",
    #     "tbl_it_name", "tbl_business_name", "tbl_it_type",
    # ]

    print("Finding the terms associated with each entity")
    relationships = []
    known_pairs = set()
    # Used only in counting
    known_entities = set()
    known_terms = set()
    
    # Iterate over the sheet's data and build our relationships
    for row in sheet_data:
        table_term = row["tbl_business_name"] + '@' + args.glossary
        column_term = row["fld_business_name"] + '@' + args.glossary
        table_name = row["tbl_it_name"]
        table_type = row["tbl_it_type"]
        column_name = row["fld_it_name"]
        column_type = row["fld_it_type"]

        known_entities.add(table_name)
        known_entities.add(column_name)

        known_terms.add(table_term)
        known_terms.add(column_term)

        table = AtlasEntity(
            name=table_name,
            qualified_name=table_name,
            typeName=table_type
        )

        column = AtlasEntity(
            name=column_name,
            qualified_name=column_name,
            typeName=column_type
        )
        
        row_pairs = [ (table, table_term), (column, column_term) ]
        
        # Create a relationship between each entity and glossary term
        for term_pair in row_pairs:
            print(term_pair)
            if term_pair in known_pairs:
                continue
            else:
                known_pairs.add(term_pair)
                pair_entity = term_pair[0]
                pair_term = term_pair[1]
                relationship = {
                    "typeName": "AtlasGlossarySemanticAssignment",
                    "attributes": {},
                    "guid": -100,
                    "end1": {
                        "typeName": "AtlasGlossaryTerm",
                        "uniqueAttributes": {
                            "qualifiedName": pair_term
                        }
                    },
                    "end2": {
                        "typeName": pair_entity.typeName,
                        "uniqueAttributes": {
                            "qualifiedName": pair_entity.get_qualified_name()
                        }
                    }
                }
                relationships.append(relationship)

    print(f"Found {len(known_terms)} terms.")
    print(f"Found {len(known_entities)} entities.")
    print(f"Found {len(relationships)} unique relationships.")

    # Upload relationship one by one
    for relationship in relationships:
        term = relationship["end1"]["uniqueAttributes"]["qualifiedName"]
        entity = relationship["end2"]["uniqueAttributes"]["qualifiedName"]
        print(f"Working on {entity}:{term}")
        try:
            results = atlas_client.upload_relationship(relationship)
            print("\tSuccess")
        except Exception as e:
            print(
                f"Exception for {term}:{entity} and was not uploaded: {e}")
